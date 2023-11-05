from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import Http404, JsonResponse, HttpResponse
from django.contrib.auth.models import User
from .models import StudentUserInfo
from .forms import GradeExcelForm, StudentExcelForm, SubjectExcelForm
from course.models import Grades, Student, Subject
from grades.models import GradesSubject
from openpyxl import load_workbook
from django.contrib.auth.base_user import BaseUserManager
import unicodedata
import xlsxwriter
import io


def generate_username(full_name):
    name = str(full_name)    
    
    name = name.split(' ')
 
    lastname = name[-1]
    firstname = name[0]
    print(name)
    while True:
        if lastname == '':
            del name[-1]
            lastname = name[-1]
        else:
            break
    
    username = '%s%s' % (firstname[0], lastname)
    if User.objects.filter(username=username).count() > 0:
        username = '%s%s' % (firstname, lastname[0])
        if User.objects.filter(username=username).count() > 0:
            
            users = User.objects.filter(username__regex=r'^%s[1-9]{1,}$' % firstname).order_by('username').values('username')                
            if len(users) > 0:
                last_number_used = map(lambda x: int(x['username'].replace(firstname,'')), users)
                last_number_used = list(last_number_used)
                last_number_used.sort()
                last_number_used = last_number_used[-1]
                number = last_number_used + 1
                username = '%s%s' % (firstname, number)
            else:
                username = '%s%s' % (firstname, 1)
    
    return username


# Create your views here.
@login_required
@user_passes_test(lambda u: u.groups.filter(name='admin').count() > 0, login_url='/accounts/login/')
def setting_grade(request):
    if request.method == 'POST':
        grade_file = GradeExcelForm(request.POST, request.FILES)
        if grade_file.is_valid():
            cd = grade_file.cleaned_data
            the_file = cd['file']
            wb = load_workbook(the_file)

            # Get sheet names
            sheet = wb['Sheet1']

            i = 2
            while True:
                s = sheet['A'+str(i)].value
                print(s)
                if s == None:
                    break
                Grades.objects.create(grade_name=s)
                i+=1

        return redirect('setting_grade_list')
    else:
        grade_excel = GradeExcelForm()
    return render(request, 'setting/add_grade.html', {'grade_excel': grade_excel})


@login_required
@user_passes_test(lambda u: u.groups.filter(name='admin').count() > 0, login_url='/accounts/login/')
def grade_list(request):
    grades = Grades.objects.all()
    return render(request,'setting/setting_grade_list.html', {'grades': grades})


@login_required
@user_passes_test(lambda u: u.groups.filter(name='admin').count() > 0, login_url='/accounts/login/')
def setting_student(request):
    
    if request.method == 'POST':
        stud_file = StudentExcelForm(request.POST, request.FILES)
        if stud_file.is_valid():
            cd = stud_file.cleaned_data
            the_file = cd['file']
            wb = load_workbook(the_file)
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output)
            cell_format = workbook.add_format({'bold': True, 'italic': True})
            sheet_list = wb.sheetnames
            for i in sheet_list:
                sheet = wb[i]
                gr = Grades.objects.get(grade_name=i)
                worksheet = workbook.add_worksheet(i)
                worksheet.set_column(0, 0, 16)
                worksheet.set_column(1, 3, 30)
                worksheet.write(0, 0, 'ID', cell_format)
                worksheet.write(0, 1, 'Full Name', cell_format)
                worksheet.write(0, 2, 'Username', cell_format)
                worksheet.write(0, 3, 'Password', cell_format)
                i = 2
                row = 1
                while True:
                    stud_id = sheet['A'+str(i)].value
                    stud_name = sheet['B'+str(i)].value

                    if stud_id == None:
                        break 

                    user_n = stud_id.split('/')
                    username = ''
                    for n in user_n:
                        username+=n 

                          
                    password = BaseUserManager().make_random_password()
                    gender = sheet['C'+str(i)].value

                    first_name = stud_name.split(' ')[0]    
                    last_n = stud_name.split(' ')
                    last_name = ''
                    del last_n[0]
                    for l in last_n:
                        last_name+=l
                        last_name+=' '

                    
                    new_user = User.objects.create_user(username=username, password=password, first_name=first_name, last_name=last_name)
                    new_user.save()
                    student = Student.objects.create(user=new_user, stud_id=stud_id, grade=gr, gender=gender)
                    StudentUserInfo.objects.create(user_id=stud_id, name=stud_name, grade=gr.grade_name, username=username, password=password)
                    worksheet.write(row, 0, stud_id)
                    worksheet.write(row, 1, stud_name)
                    worksheet.write(row, 2, username)
                    worksheet.write(row, 3, password)
                    
                    row+=1
                    i+=1                                    
            workbook.close()
            output.seek(0)
            filename = 'Registered students account info.xlsx'
            response = HttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=%s' % filename

            return response
        return HttpResponse()
    else:
        stud_excel = GradeExcelForm()
    return render(request, 'setting/add_student.html', {'stud_excel': stud_excel})


@login_required
@user_passes_test(lambda u: u.groups.filter(name='admin').count() > 0, login_url='/accounts/login/')
def setting_subject(request):
    if request.method == 'POST':
        alph = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        subject_file = SubjectExcelForm(request.POST, request.FILES)
        if subject_file.is_valid():
            cd = subject_file.cleaned_data
            the_file = cd['file']
            wb = load_workbook(the_file)

            # Get sheet names
            sheet = wb['Sheet1']

            i = 2
            while True:
                sub_name = sheet['A'+str(i)].value
                if sub_name == None:
                    break

                grades_list = str(sheet['B'+str(i)].value).split(', ')
                del grades_list[-1]
                new_sub = Subject.objects.create(title=sub_name, slug=sub_name)
                for g in grades_list:
                    get_grade = True
                    al = 0
                    while get_grade:
                        print(g+alph[al])
                        if Grades.objects.filter(grade_name=(g+alph[al])).exists():
                            gr = Grades.objects.get(grade_name=(g+alph[al]))
                            GradesSubject.objects.create(grade=gr, subject=new_sub)
                            print(gr.grade_name)
                        else:
                            get_grade = False
                        al+=1

                
                print(grades_list)
            
                
                i+=1

        return redirect('subject_list')
    else:
        subject_excel = SubjectExcelForm()
    return render(request, 'setting/add_subject.html', {'subject_excel': subject_excel})



@login_required
@user_passes_test(lambda u: u.groups.filter(name='admin').count() > 0, login_url='/accounts/login/')
def subject_list(request):
    subjects = Subject.objects.all()
    return render(request, 'setting/subject_list.html', {'subjects': subjects})
